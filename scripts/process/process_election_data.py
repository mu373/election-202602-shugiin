#!/usr/bin/env python3
"""都道府県別有権者数・投票率の整形

入力: data/raw/election/yukensha/todofuken_yukensha_touhyoritsu.xlsx
出力: data/raw/election/yukensha/r8_todofuken_yukensha.csv
"""

import pandas as pd
import openpyxl
from pathlib import Path

BASE = Path(__file__).resolve().parent.parent.parent
RAW_ELECTION = BASE / "data" / "raw" / "election" / "yukensha"
INPUT_FILE = RAW_ELECTION / "todofuken_yukensha_touhyoritsu.xlsx"
OUTPUT_FILE = RAW_ELECTION / "r8_todofuken_yukensha.csv"

def process_todofuken_yukensha():
    """Process prefecture-level voter turnout data"""

    if not INPUT_FILE.exists():
        print(f"❌ Input file not found: {INPUT_FILE}")
        print("   Run: bash scripts/download_all.sh")
        return

    print(f"Processing: {INPUT_FILE.name}")

    wb = openpyxl.load_workbook(INPUT_FILE)
    ws = wb['投票結果']

    # Extract data
    data = []
    current_block = None

    for r in range(6, 63):  # Rows 6-62
        block = ws.cell(r, 1).value
        pref = ws.cell(r, 2).value

        if block:
            current_block = block

        # Skip subtotal rows
        if not pref or pref in ['計', '合計']:
            continue

        yukensha_total = ws.cell(r, 5).value
        touhyousha_total = ws.cell(r, 8).value

        if yukensha_total and touhyousha_total:
            data.append({
                'block_name': current_block,
                'pref_name': pref,
                'yukensha_male': int(ws.cell(r, 3).value or 0),
                'yukensha_female': int(ws.cell(r, 4).value or 0),
                'yukensha_total': int(yukensha_total),
                'touhyousha_male': int(ws.cell(r, 6).value or 0),
                'touhyousha_female': int(ws.cell(r, 7).value or 0),
                'touhyousha_total': int(touhyousha_total),
                'kikensha_male': int(ws.cell(r, 9).value or 0),
                'kikensha_female': int(ws.cell(r, 10).value or 0),
                'kikensha_total': int(ws.cell(r, 11).value or 0),
            })

    df = pd.DataFrame(data)
    df['touhyou_rate'] = df['touhyousha_total'] / df['yukensha_total'] * 100

    # Add pref_code
    pref_codes = {
        '北海道': '01', '青森県': '02', '岩手県': '03', '宮城県': '04', '秋田県': '05',
        '山形県': '06', '福島県': '07', '茨城県': '08', '栃木県': '09', '群馬県': '10',
        '埼玉県': '11', '千葉県': '12', '東京都': '13', '神奈川県': '14', '新潟県': '15',
        '富山県': '16', '石川県': '17', '福井県': '18', '山梨県': '19', '長野県': '20',
        '岐阜県': '21', '静岡県': '22', '愛知県': '23', '三重県': '24', '滋賀県': '25',
        '京都府': '26', '大阪府': '27', '兵庫県': '28', '奈良県': '29', '和歌山県': '30',
        '鳥取県': '31', '島根県': '32', '岡山県': '33', '広島県': '34', '山口県': '35',
        '徳島県': '36', '香川県': '37', '愛媛県': '38', '高知県': '39', '福岡県': '40',
        '佐賀県': '41', '長崎県': '42', '熊本県': '43', '大分県': '44', '宮崎県': '45',
        '鹿児島県': '46', '沖縄県': '47'
    }
    df.insert(0, 'pref_code', df['pref_name'].map(pref_codes))

    # Reorder columns
    df = df[['pref_code', 'pref_name', 'block_name',
             'yukensha_male', 'yukensha_female', 'yukensha_total',
             'touhyousha_male', 'touhyousha_female', 'touhyousha_total',
             'kikensha_male', 'kikensha_female', 'kikensha_total',
             'touhyou_rate']]

    # Save
    df.to_csv(OUTPUT_FILE, index=False, encoding='utf-8')

    print(f"✓ Saved: {OUTPUT_FILE}")
    print(f"  {len(df)} prefectures")
    print(f"  有権者数合計: {df['yukensha_total'].sum():,} 人")
    print(f"  投票者数合計: {df['touhyousha_total'].sum():,} 人")
    print(f"  全国投票率: {df['touhyou_rate'].mean():.2f}%")

if __name__ == '__main__':
    process_todofuken_yukensha()
