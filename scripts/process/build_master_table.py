#!/usr/bin/env python3
"""マスター対応表の構築（選挙区 → 都道府県 → 比例ブロック）

入力: data/raw/gis/senkyoku_ichiran.xlsx（289選挙区リスト）
出力: data/master/district_master.csv（289行: 選挙区・都道府県・ブロック対応）
"""

import pandas as pd
import openpyxl
from pathlib import Path

# Paths
BASE = Path(__file__).resolve().parent.parent.parent
RAW = str(BASE / "data" / "raw")
MASTER = str(BASE / "data" / "master")

# 1. Build pref → block mapping (hardcoded, 47 prefectures → 11 blocks)
_PREF_BLOCK_ROWS = [
    ("01","北海道",1,"北海道"),("02","青森県",2,"東北"),("03","岩手県",2,"東北"),
    ("04","宮城県",2,"東北"),("05","秋田県",2,"東北"),("06","山形県",2,"東北"),
    ("07","福島県",2,"東北"),("08","茨城県",3,"北関東"),("09","栃木県",3,"北関東"),
    ("10","群馬県",3,"北関東"),("11","埼玉県",3,"北関東"),("12","千葉県",4,"南関東"),
    ("13","東京都",5,"東京"),("14","神奈川県",4,"南関東"),("15","新潟県",6,"北陸信越"),
    ("16","富山県",6,"北陸信越"),("17","石川県",6,"北陸信越"),("18","福井県",6,"北陸信越"),
    ("19","山梨県",7,"東海"),("20","長野県",6,"北陸信越"),("21","岐阜県",7,"東海"),
    ("22","静岡県",7,"東海"),("23","愛知県",7,"東海"),("24","三重県",7,"東海"),
    ("25","滋賀県",8,"近畿"),("26","京都府",8,"近畿"),("27","大阪府",8,"近畿"),
    ("28","兵庫県",8,"近畿"),("29","奈良県",8,"近畿"),("30","和歌山県",8,"近畿"),
    ("31","鳥取県",9,"中国"),("32","島根県",9,"中国"),("33","岡山県",9,"中国"),
    ("34","広島県",9,"中国"),("35","山口県",9,"中国"),("36","徳島県",10,"四国"),
    ("37","香川県",10,"四国"),("38","愛媛県",10,"四国"),("39","高知県",10,"四国"),
    ("40","福岡県",11,"九州"),("41","佐賀県",11,"九州"),("42","長崎県",11,"九州"),
    ("43","熊本県",11,"九州"),("44","大分県",11,"九州"),("45","宮崎県",11,"九州"),
    ("46","鹿児島県",11,"九州"),("47","沖縄県",11,"九州"),
]
pref_block = pd.DataFrame(_PREF_BLOCK_ROWS, columns=["pref_code","pref_name","block_id","block_name"])
print(f"Built {len(pref_block)} prefectures → 11 blocks")

# 2. Load G3: district list (contains pref code)
wb = openpyxl.load_workbook(f"{RAW}/gis/senkyoku_ichiran.xlsx")
ws = wb['Sheet1']

districts = []
for row in range(5, 294):  # Rows 5-293 = 289 districts
    pref_code = ws.cell(row, 1).value
    district_num = ws.cell(row, 2).value
    district_code = ws.cell(row, 3).value
    district_name = ws.cell(row, 4).value

    if pref_code is None:
        break

    districts.append({
        'pref_code': str(int(pref_code)).zfill(2),  # Convert to string with leading zero
        'district_num': int(district_num),
        'district_code': district_code,
        'district_name': district_name,
    })

df_district = pd.DataFrame(districts)
print(f"Loaded {len(df_district)} electoral districts")

# 3. Merge pref → block
df_district = df_district.merge(
    pref_block[['pref_code', 'pref_name', 'block_id', 'block_name']],
    on='pref_code',
    how='left'
)

print(f"\nDistrict → Block mapping:")
print(df_district.groupby('block_name').size())

# 4. Save district master
df_district.to_csv(f"{MASTER}/district_master.csv", index=False, encoding='utf-8')
print(f"\nSaved: {MASTER}/district_master.csv")
print(f"Columns: {list(df_district.columns)}")
print(f"\nSample:")
print(df_district.head(10))

print("\n" + "="*60)
print("NOTE: Municipality → District mapping requires election data")
print("      (E2: 市区町村別得票数) which identifies split municipalities.")
print("      This will be built in Phase 1 after downloading election data.")
